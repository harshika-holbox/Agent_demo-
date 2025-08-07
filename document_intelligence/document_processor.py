import os
import io
from typing import Tuple
import PyPDF2
import pdfplumber
from PIL import Image
import pytesseract
from docx import Document
import openpyxl
import pandas as pd

class DocumentProcessor:
    """
    Handles processing of various document formats including PDFs, images, and office documents
    """
    
    def __init__(self):
        self.supported_formats = {
            'pdf': ['.pdf'],
            'image': ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.webp'],
            'word': ['.docx', '.doc'],
            'excel': ['.xlsx', '.xls'],
            'csv': ['.csv'],
            'text': ['.txt', '.md', '.rtf']
        }
    
    def get_file_extension(self, filename: str) -> str:
        """Get file extension from filename"""
        return os.path.splitext(filename.lower())[1]
    
    def is_supported_format(self, filename: str) -> bool:
        """Check if file format is supported"""
        ext = self.get_file_extension(filename)
        return any(ext in formats for formats in self.supported_formats.values())
    
    def get_file_type(self, filename: str) -> str:
        """Get the type of file based on extension"""
        ext = self.get_file_extension(filename)
        for file_type, extensions in self.supported_formats.items():
            if ext in extensions:
                return file_type
        return 'unknown'
    
    def extract_text_from_pdf(self, file_content: bytes) -> str:
        """Extract text from PDF file"""
        try:
            # Try pdfplumber first (better for complex layouts)
            with pdfplumber.open(io.BytesIO(file_content)) as pdf:
                text = ""
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                return text.strip()
        except Exception as e:
            # Fallback to PyPDF2
            try:
                pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                return text.strip()
            except Exception as e2:
                raise Exception(f"Failed to extract text from PDF: {str(e2)}")
    
    def extract_text_from_image(self, file_content: bytes) -> str:
        """Extract text from image using OCR with enhanced error handling"""
        try:
            # Validate file content
            if len(file_content) == 0:
                raise Exception("Empty file content")
            
            print(f"Processing image: {len(file_content)} bytes")
            
            # Try to open image with better error handling
            try:
                image = Image.open(io.BytesIO(file_content))
                print(f"Image opened successfully: {image.format}, {image.mode}, {image.size}")
            except Exception as img_error:
                # Try to save raw bytes for debugging
                print(f"PIL Image.open failed: {img_error}")
                
                # Check if it's a valid image format by magic bytes
                magic_bytes = file_content[:10]
                print(f"File magic bytes: {magic_bytes}")
                
                # Try alternative approach - save to temp file
                import tempfile
                try:
                    with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
                        temp_file.write(file_content)
                        temp_file.flush()
                        image = Image.open(temp_file.name)
                        print(f"Image opened via temp file: {image.format}, {image.mode}, {image.size}")
                    os.unlink(temp_file.name)  # Clean up temp file
                except Exception as temp_error:
                    raise Exception(f"Cannot open image file. PIL error: {img_error}, Temp file error: {temp_error}")
            
            # Convert to RGB if necessary
            if image.mode != 'RGB':
                print(f"Converting image from {image.mode} to RGB")
                image = image.convert('RGB')
            
            # Check if Tesseract is available
            try:
                pytesseract.get_tesseract_version()
                print(f"Tesseract available: {pytesseract.get_tesseract_version()}")
            except Exception as tess_error:
                raise Exception(f"Tesseract OCR not available: {tess_error}")
            
            # Extract text using OCR with configuration
            print("Starting OCR text extraction...")
            
            # Use OCR configuration for better results
            custom_config = r'--oem 3 --psm 6'
            text = pytesseract.image_to_string(image, config=custom_config)
            
            extracted_text = text.strip()
            print(f"OCR completed. Extracted {len(extracted_text)} characters")
            
            if not extracted_text:
                print("Warning: No text extracted from image")
                return "[No text content detected in this image]"
            
            return extracted_text
            
        except Exception as e:
            error_msg = f"Failed to extract text from image: {str(e)}"
            print(f"Error in extract_text_from_image: {error_msg}")
            raise Exception(error_msg)
    
    def extract_text_from_word(self, file_content: bytes) -> str:
        """Extract text from Word document"""
        try:
            doc = Document(io.BytesIO(file_content))
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text.strip()
        except Exception as e:
            raise Exception(f"Failed to extract text from Word document: {str(e)}")
    
    def extract_text_from_excel(self, file_content: bytes) -> str:
        """Extract text from Excel file"""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"\n--- Sheet: {sheet_name} ---\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        text += row_text + "\n"
            
            return text.strip()
        except Exception as e:
            raise Exception(f"Failed to extract text from Excel file: {str(e)}")
    
    def extract_text_from_csv(self, file_content: bytes) -> str:
        """Extract text from CSV file"""
        try:
            # Try to read as CSV
            df = pd.read_csv(io.BytesIO(file_content))
            return df.to_string(index=False)
        except Exception as e:
            raise Exception(f"Failed to extract text from CSV file: {str(e)}")
    
    def extract_text_from_text_file(self, file_content: bytes) -> str:
        """Extract text from plain text file"""
        try:
            # Try different encodings
            encodings = ['utf-8', 'latin-1', 'cp1252']
            for encoding in encodings:
                try:
                    return file_content.decode(encoding).strip()
                except UnicodeDecodeError:
                    continue
            raise Exception("Could not decode text file with any supported encoding")
        except Exception as e:
            raise Exception(f"Failed to extract text from text file: {str(e)}")
    
    def process_file(self, file_content: bytes, filename: str) -> Tuple[str, str]:
        """
        Process a file and extract text content
        
        Returns:
            Tuple of (extracted_text, file_type)
        """
        if not self.is_supported_format(filename):
            raise Exception(f"Unsupported file format: {filename}")
        
        file_type = self.get_file_type(filename)
        
        try:
            if file_type == 'pdf':
                text = self.extract_text_from_pdf(file_content)
            elif file_type == 'image':
                text = self.extract_text_from_image(file_content)
            elif file_type == 'word':
                text = self.extract_text_from_word(file_content)
            elif file_type == 'excel':
                text = self.extract_text_from_excel(file_content)
            elif file_type == 'csv':
                text = self.extract_text_from_csv(file_content)
            elif file_type == 'text':
                text = self.extract_text_from_text_file(file_content)
            else:
                raise Exception(f"Unknown file type: {file_type}")
            
            # Special handling for images - they might not have extractable text
            if not text.strip():
                if file_type == 'image':
                    # For images, provide a descriptive message instead of failing
                    text = f"[Image file uploaded: {filename}. This appears to be a visual/graphical image with limited text content suitable for OCR extraction. The image may contain charts, diagrams, or complex layouts that cannot be easily converted to text.]"
                else:
                    raise Exception("No text content extracted from file")
            
            return text, file_type
            
        except Exception as e:
            raise Exception(f"Error processing {filename}: {str(e)}")
    
    def get_supported_formats_info(self) -> str:
        """Get information about supported file formats"""
        info = "Supported file formats:\n"
        for file_type, extensions in self.supported_formats.items():
            info += f"- {file_type.upper()}: {', '.join(extensions)}\n"
        return info 