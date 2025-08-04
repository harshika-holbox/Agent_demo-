import os
import io
import base64
import requests
from typing import Optional, Tuple, List
import PyPDF2
import pdfplumber
from PIL import Image
import pytesseract
from docx import Document
import openpyxl
import pandas as pd
import json

class DocumentProcessor:
    """
    Handles processing of various document formats including PDFs, images, and office documents
    """
    
    def __init__(self):
        self.supported_formats = {
            'pdf': ['.pdf'],
            'image': ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.webp'],
            'word': ['.docx', '.doc'],
            'excel': ['.xlsx', '.xls'],
            'csv': ['.csv'],
            'text': ['.txt', '.md', '.rtf']
        }
    
    def get_file_extension(self, filename: str) -> str:
        """Get file extension from filename"""
        return os.path.splitext(filename.lower())[1]
    
    def is_supported_format(self, filename: str) -> bool:
        """Check if file format is supported"""
        ext = self.get_file_extension(filename)
        return any(ext in formats for formats in self.supported_formats.values())
    
    def get_file_type(self, filename: str) -> str:
        """Get the type of file based on extension"""
        ext = self.get_file_extension(filename)
        for file_type, extensions in self.supported_formats.items():
            if ext in extensions:
                return file_type
        return 'unknown'
    
    def extract_text_from_pdf(self, file_content: bytes) -> str:
        """Extract text from PDF file"""
        try:
            # Try pdfplumber first (better for complex layouts)
            with pdfplumber.open(io.BytesIO(file_content)) as pdf:
                text = ""
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                return text.strip()
        except Exception as e:
            # Fallback to PyPDF2
            try:
                pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                return text.strip()
            except Exception as e2:
                raise Exception(f"Failed to extract text from PDF: {str(e2)}")
    
    def extract_text_from_image(self, file_content: bytes) -> str:
        """Extract text from image using OCR"""
        try:
            # Open image
            image = Image.open(io.BytesIO(file_content))
            
            # Convert to RGB if necessary
            if image.mode != 'RGB':
                image = image.convert('RGB')
            
            # Extract text using OCR
            text = pytesseract.image_to_string(image)
            return text.strip()
        except Exception as e:
            raise Exception(f"Failed to extract text from image: {str(e)}")
    
    def extract_text_from_word(self, file_content: bytes) -> str:
        """Extract text from Word document"""
        try:
            doc = Document(io.BytesIO(file_content))
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text.strip()
        except Exception as e:
            raise Exception(f"Failed to extract text from Word document: {str(e)}")
    
    def extract_text_from_excel(self, file_content: bytes) -> str:
        """Extract text from Excel file"""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"\n--- Sheet: {sheet_name} ---\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        text += row_text + "\n"
            
            return text.strip()
        except Exception as e:
            raise Exception(f"Failed to extract text from Excel file: {str(e)}")
    
    def extract_text_from_csv(self, file_content: bytes) -> str:
        """Extract text from CSV file"""
        try:
            # Try to read as CSV
            df = pd.read_csv(io.BytesIO(file_content))
            return df.to_string(index=False)
        except Exception as e:
            raise Exception(f"Failed to extract text from CSV file: {str(e)}")
    
    def extract_text_from_text_file(self, file_content: bytes) -> str:
        """Extract text from plain text file"""
        try:
            # Try different encodings
            encodings = ['utf-8', 'latin-1', 'cp1252']
            for encoding in encodings:
                try:
                    return file_content.decode(encoding).strip()
                except UnicodeDecodeError:
                    continue
            raise Exception("Could not decode text file with any supported encoding")
        except Exception as e:
            raise Exception(f"Failed to extract text from text file: {str(e)}")
    
    def process_file(self, file_content: bytes, filename: str) -> Tuple[str, str]:
        """
        Process a file and extract text content
        
        Returns:
            Tuple of (extracted_text, file_type)
        """
        if not self.is_supported_format(filename):
            raise Exception(f"Unsupported file format: {filename}")
        
        file_type = self.get_file_type(filename)
        
        try:
            if file_type == 'pdf':
                text = self.extract_text_from_pdf(file_content)
            elif file_type == 'image':
                text = self.extract_text_from_image(file_content)
            elif file_type == 'word':
                text = self.extract_text_from_word(file_content)
            elif file_type == 'excel':
                text = self.extract_text_from_excel(file_content)
            elif file_type == 'csv':
                text = self.extract_text_from_csv(file_content)
            elif file_type == 'text':
                text = self.extract_text_from_text_file(file_content)
            else:
                raise Exception(f"Unknown file type: {file_type}")
            
            if not text.strip():
                raise Exception("No text content extracted from file")
            
            return text, file_type
            
        except Exception as e:
            raise Exception(f"Error processing {filename}: {str(e)}")
    
    def get_supported_formats_info(self) -> str:
        """Get information about supported file formats"""
        info = "Supported file formats:\n"
        for file_type, extensions in self.supported_formats.items():
            info += f"- {file_type.upper()}: {', '.join(extensions)}\n"
        return info 